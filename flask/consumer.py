from flask import Flask
from flask import request
from flask import render_template
import image_processor as image_processor
import base64
import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Side, Border
from io import BytesIO
from flask import send_file
import os
import json
import time

app = Flask(__name__)

def apply_viridis_colormap(blurred_image:np.array) -> np.array:
    viridis_image = cv2.applyColorMap(blurred_image, cv2.COLORMAP_VIRIDIS)
    return viridis_image


def decode_base64_image(base64_image: str) -> np.array:
    try:
        # Удаление префикса, если он присутствует
        if base64_image.startswith('data:image'):
            base64_image = base64_image.split(',')[1]
        
        # Декодирование base64
        image_data = base64.b64decode(base64_image)
        # Преобразование в numpy массив
        np_array = np.frombuffer(image_data, np.uint8)
        # Декодирование изображения
        image = cv2.imdecode(np_array, cv2.IMREAD_COLOR)
        
        if image is None:
            print("Ошибка: cv2.imdecode вернул None. Возможно, данные изображения повреждены или имеют неправильный формат.")
        return image
    except Exception as e:
        print(f"Ошибка при декодировании изображения: {e}")
        return None

def apply_gaussian_blur(image: np.array) -> np.array:
    processor = image_processor.ImageProcessor(image)
    blurred_image = processor.blurGaussian((3, 3), 0, 0)
    return blurred_image

def encode_image_to_base64(viridis_image: np.array) -> str:
    _, img_encoded = cv2.imencode('.png', viridis_image)
    result = base64.b64encode(img_encoded).decode()
    return result


def remove_similar_values(lines: list[int]) -> list[int]:
    for i in range(len(lines) - 1, 0, -1):
        if abs(lines[i] - lines[i - 1]) == 0:
            lines.pop(i)
    return lines


def compute_luminance_relative_to_lines(horizontal_lines_: list, vertical_lines_: list, image: np.array) -> tuple:
    processor = image_processor.ImageProcessor(image)
    horizontal_lines = horizontal_lines_
    horizontal_lines.insert(0, 0)
    horizontal_lines.append(image.shape[0])
    horizontal_lines = list(filter(lambda x: x >= 0 and x <= image.shape[0], horizontal_lines))
    horizontal_lines = list(map(int, horizontal_lines))
    horizontal_lines = sorted(horizontal_lines)
    horizontal_lines = remove_similar_values(horizontal_lines)
    vertical_lines = vertical_lines_
    vertical_lines.insert(0, 0)
    vertical_lines.append(image.shape[1])
    vertical_lines = list(filter(lambda x: x >= 0 and x <= image.shape[1], vertical_lines))
    vertical_lines = list(map(int, vertical_lines))
    vertical_lines = sorted(vertical_lines)
    vertical_lines = remove_similar_values(vertical_lines)
    luminance_values = processor.calculateMeanRelativeToLines(vertical_lines, horizontal_lines)
    grid = (horizontal_lines, vertical_lines)
    return luminance_values, grid

def find_split_positions(image, num_splits, axis=0, alpha=0.5, threshold=128):
    """
    Находит позиции разбиения вдоль заданной оси (0 — столбцы, 1 — строки),
    учитывая составной показатель, который объединяет нормированную среднюю светимость
    и долю ярких пикселей (выше порога).
    """
    if axis == 0:
        # Анализируем столбцы
        brightness_sum = np.sum(image, axis=0)  # сумма значений яркости по столбцам
        binary_count = np.sum(image > threshold, axis=0)  # количество «ярких» пикселей по столбцам
        total_pixels = image.shape[0]  # количество пикселей в каждом столбце
    else:
        # Анализируем строки
        brightness_sum = np.sum(image, axis=1)
        binary_count = np.sum(image > threshold, axis=1)
        total_pixels = image.shape[1]
    
    # Нормируем: средняя светимость в диапазоне 0..1 (максимум – 255)
    norm_brightness = brightness_sum / (255 * total_pixels)
    # Доля ярких пикселей
    norm_binary = binary_count / total_pixels
    
    # Составной показатель
    composite = alpha * norm_brightness + (1 - alpha) * norm_binary
    
    # Кумулятивная сумма композитного показателя
    cumsum = np.cumsum(composite)
    total = cumsum[-1]
    step = total / (num_splits + 1)
    positions = []
    for i in range(1, num_splits + 1):
        pos = np.searchsorted(cumsum, i * step)
        positions.append(pos)
    return np.array(positions).tolist()

@app.route("/gaussian-blur", methods=["POST"])
def blur_image():
    if request.method == "POST":
        # base64image = request.args.get("image", default=None, type=str)
        base64image = request.json["image"]
        image = decode_base64_image(base64image)
        blurred_image = apply_gaussian_blur(image)
        viridis_image = apply_viridis_colormap(blurred_image)
        result = encode_image_to_base64(viridis_image)
        return {"image": result}, 200

@app.route("/mean-luminance", methods=["POST"])
def mean_luminance():
    if request.method == "POST":
        base64image = request.json["image"]
        image = decode_base64_image(base64image)
        # vertical_lines = request.args.get("vertical_lines", default=None, type=list)
        # horizontal_lines = request.args.get("horizontal_lines", default=None, type=list)
        vertical_lines = request.json["vertical_lines"]
        horizontal_lines = request.json["horizontal_lines"]
        #image = apply_gaussian_blur(image)
        luminance_values, grid = compute_luminance_relative_to_lines(horizontal_lines, vertical_lines, image)
        return {"luminance": luminance_values.tolist(), "grid": grid}, 200

@app.route("/adaptive-grid", methods=["POST"])
def adaptive_grid():
    if request.method == "POST":
        base64image = request.json["image"]
        image = decode_base64_image(base64image)
        gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        
        num_vertical = int(request.json.get("num_vertical", 10))
        num_horizontal = int(request.json.get("num_horizontal", 5))
        alpha = float(request.json.get("alpha", 0.5))
        threshold = int(request.json.get("threshold", 128))
        
        # Вычисляем оптимальные позиции
        vertical_lines = find_split_positions(gray_image, num_vertical, axis=0, 
                                             alpha=alpha, threshold=threshold)
        horizontal_lines = find_split_positions(gray_image, num_horizontal, axis=1, 
                                               alpha=alpha, threshold=threshold)
        
        return {"vertical_lines": vertical_lines, "horizontal_lines": horizontal_lines}, 200

@app.route("/export-excel", methods=["POST"])
def export_excel():
    if request.method == "POST":
        try:
            # Get data from request
            data = request.json
            table_data = data.get("tableData", [])
            selected_cells = data.get("selectedCells", [])
            category_averages = data.get("categoryAverages", [])
            
            # Create a new workbook and select the active worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Данные светимости"
            
            # Define styles
            header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            corner_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            average_fill = PatternFill(start_color="E6F7FF", end_color="E6F7FF", fill_type="solid")
            overall_label_fill = PatternFill(start_color="F0F7FF", end_color="F0F7FF", fill_type="solid")
            overall_value_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            bold_font = Font(bold=True)
            blue_font = Font(bold=True, color="0056B3")
            
            # Write data to the worksheet
            for row_idx, row in enumerate(table_data):
                for col_idx, cell_value in enumerate(row):
                    excel_cell = ws.cell(row=row_idx+1, column=col_idx+1)
                    
                    # Handle numeric values
                    if isinstance(cell_value, (int, float)) or (isinstance(cell_value, str) and cell_value.replace('.', '', 1).isdigit()):
                        try:
                            excel_cell.value = float(cell_value)
                            # Format numbers to 3 decimal places to match web UI
                            if isinstance(excel_cell.value, float):
                                excel_cell.number_format = '0.000'
                        except (ValueError, TypeError):
                            excel_cell.value = cell_value
                    else:
                        excel_cell.value = cell_value
                    
                    # Apply header styling
                    if row_idx == 0:
                        excel_cell.fill = header_fill
                        excel_cell.font = bold_font
                        if col_idx == 0:
                            excel_cell.fill = corner_fill
                    elif col_idx == 0:
                        excel_cell.fill = header_fill
                        excel_cell.font = bold_font
                    
                    # Apply average styling to category columns
                    if row_idx > 0 and col_idx >= len(row) - len(category_averages):
                        category_idx = col_idx - (len(row) - len(category_averages))
                        if category_idx >= 0 and category_idx < len(category_averages):
                            cat_color = category_averages[category_idx].get('color', '#E6F7FF')
                            cat_color = cat_color.lstrip('#')
                            # Apply lighter shade for average cells
                            excel_cell.fill = PatternFill(start_color="33"+cat_color, end_color="33"+cat_color, fill_type="solid")
                            excel_cell.font = bold_font
                    
                    # Apply overall average row styling - for the bottom row
                    if row_idx == len(table_data) - 1:
                        excel_cell.font = bold_font
                        if col_idx == 0:  # Overall average label
                            excel_cell.fill = overall_label_fill
                            # Add a top border to match the web UI's division line
                            thin = Side(border_style="thin", color="007BFF")
                            excel_cell.border = Border(top=thin)
                        else:
                            # Add a top border to all cells in the bottom row
                            thin = Side(border_style="thin", color="007BFF")
                            excel_cell.border = Border(top=thin)
            
            # Create a mapping of cells to their categories
            cell_categories = {}
            # Group selected cells by position
            for cell_info in selected_cells:
                row = cell_info.get('row', 0)
                col = cell_info.get('col', 0)
                category_id = cell_info.get('categoryId')
                
                key = f"{row},{col}"
                if key not in cell_categories:
                    cell_categories[key] = []
                
                if category_id:
                    cell_categories[key].append(category_id)
            
            # Apply cell colors based on categories
            category_colors = {}
            for category in category_averages:
                category_colors[category['id']] = category['color']
            
            for key, cat_ids in cell_categories.items():
                row, col = map(int, key.split(','))
                excel_row = row + 1  # Excel is 1-indexed
                excel_col = col + 1
                
                # Skip if only one category (simple case)
                if len(cat_ids) == 1:
                    category_id = cat_ids[0]
                    if category_id and category_id in category_colors:
                        color = category_colors[category_id]
                        # Convert HTML color to Excel color
                        hex_color = color.lstrip('#')
                        if len(hex_color) == 6:
                            # Add 77 (about 50% opacity) for semi-transparent effect
                            excel_color = "77" + hex_color
                            fill = PatternFill(start_color=excel_color, end_color=excel_color, fill_type="solid")
                            ws.cell(row=excel_row+1, column=excel_col+1).fill = fill
                else:
                    # For multiple categories, we can't do gradients in Excel easily
                    # So we'll use a pattern fill with the first category color
                    if cat_ids and cat_ids[0] in category_colors:
                        color = category_colors[cat_ids[0]]
                        hex_color = color.lstrip('#')
                        if len(hex_color) == 6:
                            excel_color = "77" + hex_color
                            fill = PatternFill(start_color=excel_color, end_color=excel_color, fill_type="solid")
                            cell = ws.cell(row=excel_row+1, column=excel_col+1)
                            cell.fill = fill
                            
                            # Add a comment listing all categories
                            if len(cat_ids) > 1:
                                from openpyxl.comments import Comment
                                cat_names = [next((cat['name'] for cat in category_averages if cat['id'] == cat_id), cat_id) 
                                            for cat_id in cat_ids]
                                comment = Comment(f"Ячейка содержит несколько категорий: {', '.join(cat_names)}", "Excel Export")
                                cell.comment = comment
            
            # Add category averages table below main table
            if category_averages:
                # Add a blank row
                next_row = len(table_data) + 3
                
                # Add header
                ws.cell(row=next_row, column=1).value = "Статистика по категориям"
                ws.cell(row=next_row, column=1).font = Font(bold=True, size=14)
                ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=4)
                
                # Add table headers
                next_row += 1
                headers = ["Категория", "Цвет", "Кол-во ячеек", "Среднее значение"]
                for i, header in enumerate(headers):
                    cell = ws.cell(row=next_row, column=i+1)
                    cell.value = header
                    cell.font = bold_font
                    cell.fill = header_fill
                
                # Add category data
                for category in category_averages:
                    next_row += 1
                    ws.cell(row=next_row, column=1).value = category.get('name', 'Неизвестно')
                    
                    # Set color cell
                    color_cell = ws.cell(row=next_row, column=2)
                    color_cell.value = " "  # space to ensure the cell has content
                    color_hex = category.get('color', '#CCCCCC').lstrip('#')
                    color_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                    color_cell.fill = color_fill
                    
                    ws.cell(row=next_row, column=3).value = category.get('count', 0)
                    ws.cell(row=next_row, column=4).value = category.get('average', 'Н/Д')
                    # Format average values to 3 decimal places
                    if category.get('average') not in ('Н/Д', 'N/A'):
                        ws.cell(row=next_row, column=4).number_format = '0.000'
            
            # Set column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width
            
            # Center align all cells
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Save to a BytesIO object
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Generate a unique filename with Russian text
            filename = f"данные_светимости_{int(time.time())}.xlsx"
            
            # Send the file as an attachment
            return send_file(
                output,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=filename
            )
            
        except Exception as e:
            print(f"Ошибка при создании Excel файла: {e}")
            return {"error": str(e)}, 500

@app.route("/", methods=["GET"])
def index():
    return render_template("index.html"), 200


if __name__ == "__main__":
    app.run(host="localhost", port=8000)